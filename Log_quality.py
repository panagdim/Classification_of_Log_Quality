# GUI 
import pandas as pd
import math
import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter import ttk

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg


# =========================
# UTILITIES
# =========================

def clamp(value, min_val, max_val):
    return max(min_val, min(max_val, value))


def to_bool(val):
    return str(val).strip().lower() in ["true", "1", "yes"]


def calculate_volume(length, diameter_cm):
    radius = (diameter_cm / 100) / 2
    return round(math.pi * radius ** 2 * length, 4)


# =========================
# MODEL
# =========================

def evaluate_log(length, diameter, knots, cracks,
                 has_forking, has_bend, has_disease):

    no_defects = (
        knots == 0 and cracks == 0 and
        not has_forking and not has_bend and not has_disease
    )

    if no_defects and length >= 7 and diameter >= 80:
        return 95, "High"

    length_score = 45 * (1 - math.exp(-0.4 * (length - 2)))
    diameter_score = 45 * (diameter / 150) ** 0.85
    size_score = length_score + diameter_score

    penalty = (
        knots * 1.8 +
        cracks * (2.2 + 0.25 * length) +
        (8 if has_forking else 0) +
        (6 if has_bend else 0) +
        (12 if has_disease else 0)
    )

    interaction = 0
    if cracks > 0 and has_disease:
        interaction += 6
    if knots > 5 and cracks > 3:
        interaction += 5
    if length > 7 and has_bend:
        interaction += 4
    if has_forking:
        interaction += length * 0.4

    score = size_score - penalty - interaction
    if no_defects:
        score += 10

    score = max(0, min(100, score))

    if score >= 80:
        return score, "High"
    elif score >= 60:
        return score, "Medium"
    elif score >= 40:
        return score, "Low"
    else:
        return score, "Very Low"


# =========================
# EXCEL FORMATTING
# =========================

def format_excel(path):
    wb = load_workbook(path)
    ws = wb.active

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)

        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = max_len + 2

    colors = {
        "High": "2ECC71",
        "Medium": "F1C40F",
        "Low": "E67E22",
        "Very Low": "E74C3C"
    }

    quality_col = None
    for cell in ws[1]:
        if cell.value == "Quality":
            quality_col = cell.column_letter

    for row in range(2, ws.max_row + 1):
        cell = ws[f"{quality_col}{row}"]
        if cell.value in colors:
            cell.fill = PatternFill(
                start_color=colors[cell.value],
                end_color=colors[cell.value],
                fill_type="solid"
            )

    wb.save(path)


# =========================
# GUI APP
# =========================

class ForestApp:
    def __init__(self, root):
        self.root = root
        self.root.title("🌲 Logify")
        self.root.geometry("1000x750")

        self.file_path = None
        self.df = None

        tk.Label(
            root,
            text="🌲🪵 Dashboard 🪵🌲",
            font=("Arial", 16, "bold"),
            fg="dark green"
        ).pack(pady=10)

        tk.Button(root, text="📂 Load Excel", command=self.load_file).pack(pady=5)
        tk.Button(root, text="⚙ Run Evaluation", command=self.run).pack(pady=5)
        tk.Button(root, text="💾 Save & Exit", command=self.save).pack(pady=5)

        # TABLE
        table_frame = tk.Frame(root)
        table_frame.pack(expand=True, fill="both")

        self.tree_scroll = tk.Scrollbar(table_frame)
        self.tree_scroll.pack(side="right", fill="y")

        self.tree = ttk.Treeview(table_frame, yscrollcommand=self.tree_scroll.set)
        self.tree.pack(expand=True, fill="both")

        self.tree_scroll.config(command=self.tree.yview)

        # CHART
        self.chart_frame = tk.Frame(root)
        self.chart_frame.pack(expand=True, fill="both")

    # -------------------------
    def load_file(self):
        self.file_path = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )

        if self.file_path:
            self.df = pd.read_excel(self.file_path)
            messagebox.showinfo("Loaded", "File loaded successfully")

    # -------------------------
    def run(self):
        if self.df is None:
            messagebox.showerror("Error", "Load file first")
            return

        df = self.df.copy()

        df["Count_Knots"] = df["Count_Knots"].fillna(0)
        df["Count_Cracks"] = df["Count_Cracks"].fillna(0)

        df["Has_Forking"] = df["Has_Forking"].apply(to_bool)
        df["Has_Bend"] = df["Has_Bend"].apply(to_bool)
        df["Has_Disease"] = df["Has_Disease"].apply(to_bool)

        scores, qualities, volumes = [], [], []

        for _, row in df.iterrows():
            length = clamp(row["Length_m"], 2, 10)
            diameter = clamp(row["Diameter_cm"], 10, 150)

            volume = calculate_volume(length, diameter)

            score, quality = evaluate_log(
                length, diameter,
                int(row["Count_Knots"]),
                int(row["Count_Cracks"]),
                row["Has_Forking"],
                row["Has_Bend"],
                row["Has_Disease"]
            )

            scores.append(score)
            qualities.append(quality)
            volumes.append(volume)

        df["Volume_m3"] = volumes
        df["Score"] = [f"{s:.1f} / 100" for s in scores]
        df["Quality"] = qualities

        self.df = df

        self.show_table(df)
        self.show_chart(df)

        messagebox.showinfo("Done", "Evaluation complete!")

    # -------------------------
    def show_table(self, df):
        self.tree.delete(*self.tree.get_children())

        cols = list(df.columns)
        self.tree["columns"] = cols
        self.tree["show"] = "headings"

        for col in cols:
            self.tree.heading(col, text=col, anchor="center")
            self.tree.column(col, width=120, anchor="center")

        for _, row in df.iterrows():
            self.tree.insert("", "end", values=list(row))

    # -------------------------
    # FIXED COLORS + LEGEND
    # -------------------------

    def show_chart(self, df):
        for widget in self.chart_frame.winfo_children():
            widget.destroy()

        counts = df["Quality"].value_counts()

        # EXACT COLOR SYSTEM (NO GREY EVER)
        color_map = {
            "High": "#2ECC71",
            "Medium": "#F1C40F",
            "Low": "#E67E22",
            "Very Low": "#E74C3C"
        }

        labels = list(counts.index)
        sizes = list(counts.values)

        # STRICT mapping (no fallback → no grey bug)
        colors = [color_map[label] for label in labels]

        fig, ax = plt.subplots()
        ax.pie(sizes, labels=labels, autopct="%1.1f%%", colors=colors)
        #ax.set_title("🌲 Log Quality Breakdown")

        canvas = FigureCanvasTkAgg(fig, master=self.chart_frame)
        canvas.draw()
        canvas.get_tk_widget().pack()

    # =========================
    # LEGEND (VISUAL PANEL)
    # =========================

        #legend_frame = tk.Frame(self.chart_frame)
        #legend_frame.pack(pady=10)

        #legend_items = [
        #    ("High", "#2ECC71", "🟢"),
        #    ("Medium", "#F1C40F", "🟡"),
        #    ("Low", "#E67E22", "🟠"),
        #    ("Very Low", "#E74C3C", "🔴")
        #]

        #for name, color, emoji in legend_items:
        #    row = tk.Frame(legend_frame)
        #    row.pack(anchor="w")

        #    tk.Label(row, text="⬤", fg=color, font=("Arial", 14, "bold")).pack(side="left")
        #    tk.Label(row, text=f" {emoji} {name}", font=("Arial", 11)).pack(side="left")

    # -------------------------
    def save(self):
        if self.df is None:
            messagebox.showerror("Error", "No data")
            return

        output_path = self.file_path.replace(".xlsx", "_evaluated.xlsx")
        self.df.to_excel(output_path, index=False)

        format_excel(output_path)

        messagebox.showinfo("Saved", "File saved successfully")

        self.root.destroy()


# =========================
# RUN APP
# =========================

if __name__ == "__main__":
    root = tk.Tk()
    app = ForestApp(root)
    root.mainloop()
